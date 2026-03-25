"""
Semgrep Results Extractor (Fixed)
==================================
Place this script in your project folder alongside the Semgrep JSON output files.

Expected file naming: scenario{N}_{tool}.json (e.g., scenario10_deepseek.json)

Run: python extract_semgrep.py
Output: semgrep_results.xlsx
"""

import json
import os
import re
import glob
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCENARIOS = {
    1: "User authentication",
    2: "SQL query with user input",
    3: "Input validation",
    4: "File upload handling",
    5: "JWT implementation",
    6: "CORS configuration",
    7: "Password reset flow",
    8: "Role-based access control",
    9: "Rate limiting and error handling",
    10: "NoSQL injection",
}

TOOLS = ["deepseek", "chatgpt", "claude"]

def parse_semgrep_json(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    findings = []
    for r in data.get('results', []):
        meta = r.get('extra', {}).get('metadata', {})
        cwes = meta.get('cwe', [])
        cwe = cwes[0] if cwes else 'N/A'
        owasp_list = meta.get('owasp', [])
        owasp_2021 = 'N/A'
        for o in owasp_list:
            if '2021' in o:
                owasp_2021 = o
                break
        if owasp_2021 == 'N/A' and owasp_list:
            owasp_2021 = owasp_list[0]
        vuln_class = meta.get('vulnerability_class', [])
        vuln_class_str = vuln_class[0] if vuln_class else 'N/A'
        findings.append({
            'rule_id': r.get('check_id', 'N/A').split('.')[-1],
            'full_rule_id': r.get('check_id', 'N/A'),
            'severity': r.get('extra', {}).get('severity', 'N/A'),
            'message': r.get('extra', {}).get('message', 'N/A')[:100],
            'cwe': cwe,
            'owasp': owasp_2021,
            'vulnerability_class': vuln_class_str,
            'likelihood': meta.get('likelihood', 'N/A'),
            'impact': meta.get('impact', 'N/A'),
            'confidence': meta.get('confidence', 'N/A'),
            'line_start': r.get('start', {}).get('line', 'N/A'),
            'line_end': r.get('end', {}).get('line', 'N/A'),
            'code_snippet': r.get('extra', {}).get('lines', '')[:150],
        })
    return findings

def parse_filename(filename):
    name = filename.lower().replace('.json', '')
    scenario_match = re.search(r'scenario(\d+)', name)
    scenario_num = int(scenario_match.group(1)) if scenario_match else None
    tool = 'Unknown'
    for t in TOOLS:
        if t in name:
            tool = t.capitalize()
            break
    scenario_label = f"S{scenario_num}: {SCENARIOS.get(scenario_num, 'Unknown')}" if scenario_num else 'Unknown'
    return scenario_label, tool, scenario_num

def create_excel(all_results):
    wb = Workbook()
    hf = PatternFill('solid', fgColor='4472C4')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    nf = Font(name='Arial', size=9)
    bf = Font(name='Arial', size=9, bold=True)
    tb = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    sev_colors = {
        'CRITICAL': PatternFill('solid', fgColor='FFC7CE'),
        'ERROR': PatternFill('solid', fgColor='FFC7CE'),
        'WARNING': PatternFill('solid', fgColor='FFEB9C'),
        'INFO': PatternFill('solid', fgColor='D6E4F0'),
    }

    def sort_key(fn):
        _, _, sn = parse_filename(fn)
        return (sn or 0, parse_filename(fn)[1])

    # Sheet 1: All Findings
    ws = wb.active
    ws.title = "All Findings"
    headers = ['File', 'Scenario', 'Tool', '#', 'Severity', 'Vulnerability',
               'CWE', 'OWASP 2021', 'Impact', 'Likelihood', 'Confidence', 'Line', 'Code Snippet']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfont, Alignment(horizontal='center', vertical='center', wrap_text=True), tb
    rn = 2
    for fn in sorted(all_results.keys(), key=sort_key):
        sl, tool, _ = parse_filename(fn)
        for i, f in enumerate(all_results[fn], 1):
            d = [fn, sl, tool, i, f['severity'], f['message'], f['cwe'], f['owasp'],
                 f['impact'], f['likelihood'], f['confidence'], f['line_start'], f['code_snippet']]
            sf = sev_colors.get(f['severity'], PatternFill())
            for col, val in enumerate(d, 1):
                c = ws.cell(row=rn, column=col, value=val)
                c.font, c.border, c.fill = nf, tb, sf
                c.alignment = Alignment(wrap_text=(col in [6,7,13]), vertical='center')
            rn += 1
    for c, w in zip('ABCDEFGHIJKLM', [25,25,12,5,11,40,30,28,10,12,12,8,45]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:M{rn-1}"

    # Sheet 2: Summary by File
    ws2 = wb.create_sheet("Summary by File")
    h2 = ['File', 'Scenario', 'Tool', 'Critical', 'Error', 'Warning', 'Info', 'Total']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfont, Alignment(horizontal='center'), tb
    rn = 2
    for fn in sorted(all_results.keys(), key=sort_key):
        sl, tool, _ = parse_filename(fn)
        counts = {'CRITICAL':0, 'ERROR':0, 'WARNING':0, 'INFO':0}
        for f in all_results[fn]:
            if f['severity'] in counts: counts[f['severity']] += 1
        total = sum(counts.values())
        for col, val in enumerate([fn, sl, tool, counts['CRITICAL'], counts['ERROR'], counts['WARNING'], counts['INFO'], total], 1):
            c = ws2.cell(row=rn, column=col, value=val)
            c.font, c.border = nf, tb
            c.alignment = Alignment(horizontal='center' if col > 2 else 'left')
        rn += 1
    for c, w in zip('ABCDEFGH', [30,28,12,10,10,10,10,10]):
        ws2.column_dimensions[c].width = w
    ws2.freeze_panes = 'A2'

    # Sheet 3: Comparison by Scenario
    ws3 = wb.create_sheet("Comparison by Scenario")
    h3 = ['Scenario', 'ChatGPT Total', 'Claude Total', 'DeepSeek Total',
           'ChatGPT Critical', 'Claude Critical', 'DeepSeek Critical']
    for col, h in enumerate(h3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfont, Alignment(horizontal='center', wrap_text=True), tb
    rn = 2
    for sn in sorted(SCENARIOS.keys()):
        sl = f"S{sn}: {SCENARIOS[sn]}"
        rd = [sl]
        for tn in ['Chatgpt', 'Claude', 'Deepseek']:
            t = sum(len(f) for fn, f in all_results.items() if parse_filename(fn)[0] == sl and parse_filename(fn)[1] == tn)
            rd.append(t)
        for tn in ['Chatgpt', 'Claude', 'Deepseek']:
            cr = sum(sum(1 for f in fs if f['severity']=='CRITICAL') for fn, fs in all_results.items() if parse_filename(fn)[0]==sl and parse_filename(fn)[1]==tn)
            rd.append(cr)
        for col, val in enumerate(rd, 1):
            c = ws3.cell(row=rn, column=col, value=val)
            c.font, c.border = nf, tb
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        rn += 1
    rd = ['TOTAL']
    for tn in ['Chatgpt', 'Claude', 'Deepseek']:
        rd.append(sum(len(f) for fn, f in all_results.items() if parse_filename(fn)[1]==tn))
    for tn in ['Chatgpt', 'Claude', 'Deepseek']:
        rd.append(sum(sum(1 for f in fs if f['severity']=='CRITICAL') for fn, fs in all_results.items() if parse_filename(fn)[1]==tn))
    for col, val in enumerate(rd, 1):
        c = ws3.cell(row=rn, column=col, value=val)
        c.font, c.border = bf, tb
        c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
    ws3.column_dimensions['A'].width = 32
    for i in range(2, 8):
        ws3.column_dimensions[chr(64+i)].width = 16
    ws3.freeze_panes = 'A2'

    # Sheet 4: OWASP Distribution
    ws4 = wb.create_sheet("OWASP Distribution")
    owasp_c = {}
    for fn, findings in all_results.items():
        _, tool, _ = parse_filename(fn)
        for f in findings:
            o = f['owasp']
            if o not in owasp_c: owasp_c[o] = {}
            owasp_c[o][tool] = owasp_c[o].get(tool, 0) + 1
    h4 = ['OWASP Category', 'ChatGPT', 'Claude', 'DeepSeek', 'Total']
    for col, h in enumerate(h4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfont, Alignment(horizontal='center'), tb
    rn = 2
    for o in sorted(owasp_c.keys()):
        ts = owasp_c[o]
        cg, cl, ds = ts.get('Chatgpt',0), ts.get('Claude',0), ts.get('Deepseek',0)
        for col, val in enumerate([o, cg, cl, ds, cg+cl+ds], 1):
            c = ws4.cell(row=rn, column=col, value=val)
            c.font, c.border = nf, tb
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        rn += 1
    ws4.column_dimensions['A'].width = 45
    for x in 'BCDE': ws4.column_dimensions[x].width = 12
    ws4.freeze_panes = 'A2'

    # Sheet 5: CWE Distribution
    ws5 = wb.create_sheet("CWE Distribution")
    cwe_c = {}
    for fn, findings in all_results.items():
        _, tool, _ = parse_filename(fn)
        for f in findings:
            cw = f['cwe']
            if cw not in cwe_c: cwe_c[cw] = {}
            cwe_c[cw][tool] = cwe_c[cw].get(tool, 0) + 1
    h5 = ['CWE', 'ChatGPT', 'Claude', 'DeepSeek', 'Total']
    for col, h in enumerate(h5, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfont, Alignment(horizontal='center'), tb
    rn = 2
    for cw in sorted(cwe_c.keys()):
        ts = cwe_c[cw]
        cg, cl, ds = ts.get('Chatgpt',0), ts.get('Claude',0), ts.get('Deepseek',0)
        for col, val in enumerate([cw, cg, cl, ds, cg+cl+ds], 1):
            c = ws5.cell(row=rn, column=col, value=val)
            c.font, c.border = nf, tb
            c.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        rn += 1
    ws5.column_dimensions['A'].width = 65
    for x in 'BCDE': ws5.column_dimensions[x].width = 12
    ws5.freeze_panes = 'A2'

    return wb

def main():
    json_files = glob.glob('scenario*_*.json')
    if not json_files:
        json_files = glob.glob('*.json')
    if not json_files:
        print("No JSON files found.")
        sys.exit(1)
    print(f"Found {len(json_files)} JSON file(s):")
    all_results = {}
    total = 0
    for fp in sorted(json_files):
        try:
            findings = parse_semgrep_json(fp)
            all_results[fp] = findings
            total += len(findings)
            sl, tool, _ = parse_filename(fp)
            print(f"  {fp}: {len(findings)} findings ({sl}, {tool})")
        except Exception as e:
            print(f"  {fp}: ERROR - {e}")
    wb = create_excel(all_results)
    wb.save('semgrep_results.xlsx')
    print(f"\nTotal: {total} findings across {len(all_results)} files")
    print(f"Saved to: semgrep_results.xlsx")

if __name__ == '__main__':
    main()